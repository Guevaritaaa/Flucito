"""
Genera/actualiza la BASE DE ENTRADAS DE ALMACÉN: un solo Excel acumulativo
(no uno por factura). Cada corrida agrega lo nuevo y se sobreescribe el
mismo archivo. Emparejamiento XML <-> pdf/txt de apoyo por folio real.
Modo borrador: sin diccionario de mapeo interno ni factores de precio todavía.
"""

from __future__ import annotations

import os
import re

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.almacen.apoyo import buscar_dato, obtener_apoyo_por_folio
from app.services.almacen.extractor import CARPETA_DATOS, extraer_conceptos, leer_meta
from app.services.almacen.resumen import construir_resumen, guardar_resumen_json


COLUMNAS_ASPEL = [
    "ESTATUS", "Clave Artículo", "TIPO ELE", "DESCRIPCION", "Descripción CFDI",
    "Unidad de entrada", "Unidad de salida", "Peso nc", "Línea", "Clave SAT",
    "Clave unidad", "Con serie", "Con lote", "Con pedimento", "Tipo de costeo",
    "CLAVE ESQUEMA", "PROVEEDOR", "MONEDA", "PRECIO COMPRA", "PUBLICO",
    "MINIMO", "LIQUIDACION", "MOSTRADOR", "MAYOREO", "DISTRIBUIDOR",
    "cero", "Existencias", "Fecha de última compra",
]

TITULO = "ENTRADAS DE ALMACÉN"
COLOR_HEADER = "1F3864"  # azul marino, como tu plantilla
NOMBRE_ARCHIVO_BASE = "BASE_ENTRADAS_ALMACEN.xlsx"
NOMBRE_ARCHIVO_RESUMEN = "RESUMEN_ENTRADAS_ALMACEN.json"

# clave para no duplicar el mismo producto si vuelves a correr el script
# sobre los mismos XML (borrador: puede afinarse cuando haya UUID/folio por fila)
CLAVES_DEDUPE = ["Clave Artículo", "Fecha de última compra", "PROVEEDOR"]

PATRON_FECHA_CFDI = re.compile(r"^(\d{4})-(\d{2})-(\d{2})")


def _formatea_fecha(fecha_iso: str) -> str | None:
    """CFDI 4.0 siempre trae Fecha como AAAA-MM-DDTHH:MM:SS -> la pasamos a DD-MM-AAAA."""
    if not fecha_iso:
        return None
    m = PATRON_FECHA_CFDI.match(fecha_iso)
    if not m:
        return fecha_iso  # no debería pasar en un CFDI válido, pero no truena
    anio, mes, dia = m.groups()
    return f"{dia}-{mes}-{anio}"


def _fila_desde_concepto(c: dict, dato: dict | None) -> dict:
    clave_articulo = dato["clave_articulo"] if dato else c["no_identificacion"]
    linea = dato["linea"] if dato else None
    # descripción corta: la del pdf/txt si la encontramos, si no, cae a la del XML
    descripcion_corta = (dato.get("descripcion_corta") if dato else None) or c["descripcion"]

    return {
        "ESTATUS": "A",
        "Clave Artículo": clave_articulo,
        "TIPO ELE": "P",
        "DESCRIPCION": descripcion_corta,
        "Descripción CFDI": c["descripcion"],
        "Unidad de entrada": "pza",
        "Unidad de salida": "pza",
        "Peso nc": None,
        "Línea": linea,
        "Clave SAT": c["clave_prod_serv"],
        "Clave unidad": c["clave_unidad"],
        "Con serie": "N",
        "Con lote": "N",
        "Con pedimento": "N",
        "Tipo de costeo": "P",
        "CLAVE ESQUEMA": 1,
        "PROVEEDOR": c["proveedor_nombre"],
        "MONEDA": "1",
        "PRECIO COMPRA": c["valor_unitario"],
        "PUBLICO": None,
        "MINIMO": None,
        "LIQUIDACION": None,
        "MOSTRADOR": None,
        "MAYOREO": None,
        "DISTRIBUIDOR": None,
        "cero": "0",
        "Existencias": None,
        "Fecha de última compra": _formatea_fecha(c["fecha_compra"]),
    }


def procesar_xml(ruta_xml: str, carpeta: str) -> pd.DataFrame:
    """Un XML -> DataFrame con sus productos (todavía no escribe a disco)."""
    conceptos = extraer_conceptos(ruta_xml)
    meta = leer_meta(ruta_xml)

    apoyo = []
    if meta["folio"] is not None and meta["year"]:
        apoyo = obtener_apoyo_por_folio(carpeta, meta["year"], meta["folio"])

    datos = [
        buscar_dato(apoyo, c["no_identificacion"]) or buscar_dato(apoyo, c["codigo_secundario"])
        for c in conceptos
    ]

    # el proveedor a veces usa un formato de código totalmente distinto en el
    # pdf/txt (visto con Universal Fittings: "UQ62-DOT-06" en xml vs "D62-06"
    # en pdf). Si el match por código falló para TODO el archivo y el total
    # de filas coincide, se asume mismo orden en ambos y se empareja por posición.
    if apoyo and all(d is None for d in datos) and len(apoyo) == len(conceptos):
        print(f"[{os.path.basename(ruta_xml)}] match por código falló para todo, "
              f"uso orden posicional ({len(conceptos)} filas en ambos)")
        datos = apoyo

    filas = [_fila_desde_concepto(c, d) for c, d in zip(conceptos, datos)]
    df = pd.DataFrame(filas, columns=COLUMNAS_ASPEL)
    print(f"[{os.path.basename(ruta_xml)}] {len(df)} productos "
          f"(apoyo {'encontrado' if apoyo else 'NO encontrado'})")
    return df


def _cargar_base_existente(ruta: str) -> pd.DataFrame:
    if not os.path.exists(ruta):
        return pd.DataFrame(columns=COLUMNAS_ASPEL)
    # fila 1 = título (merged), fila 2 = encabezados, datos desde fila 3
    return pd.read_excel(ruta, header=1)


def _aplicar_estilo(ruta: str, n_columnas: int) -> None:
    wb = openpyxl.load_workbook(ruta)
    ws = wb.active
    ws.insert_rows(1)  # deja espacio para el título arriba de los encabezados

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_columnas)
    celda_titulo = ws.cell(row=1, column=1, value=TITULO)
    celda_titulo.font = Font(bold=True, color="FFFFFF", size=14)
    celda_titulo.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for col in range(1, n_columnas + 1):
        celda = ws.cell(row=2, column=col)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        celda.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"

    for col_idx in range(1, n_columnas + 1):
        letra = ws.cell(row=2, column=col_idx).column_letter
        max_len = max(
            len(str(ws.cell(row=2, column=col_idx).value or "")),
            max((len(str(ws.cell(row=r, column=col_idx).value or ""))
                 for r in range(3, ws.max_row + 1)), default=0),
        )
        ws.column_dimensions[letra].width = min(max(max_len + 2, 10), 40)

    wb.save(ruta)


def guardar_en_base_acumulada(df_nuevo: pd.DataFrame, carpeta: str = CARPETA_DATOS) -> str:
    os.makedirs(carpeta, exist_ok=True)
    ruta = os.path.join(carpeta, NOMBRE_ARCHIVO_BASE)
    existente = _cargar_base_existente(ruta)

    combinado = pd.concat([existente, df_nuevo], ignore_index=True)
    antes = len(combinado)
    combinado.drop_duplicates(subset=CLAVES_DEDUPE, keep="last", inplace=True)
    duplicados_ignorados = antes - len(combinado)
    if duplicados_ignorados:
        print(f"({duplicados_ignorados} fila(s) duplicada(s) ignorada(s))")

    combinado.to_excel(ruta, index=False)
    _aplicar_estilo(ruta, len(COLUMNAS_ASPEL))
    ruta_json = os.path.join(carpeta, NOMBRE_ARCHIVO_RESUMEN)
    guardar_resumen_json(
        construir_resumen(
            filas_nuevas=df_nuevo,
            filas_acumuladas=combinado,
            duplicados_ignorados=duplicados_ignorados,
            nombre_excel=NOMBRE_ARCHIVO_BASE,
            nombre_json=NOMBRE_ARCHIVO_RESUMEN,
        ),
        ruta_json,
    )
    print(f"-> {ruta}  ({len(combinado)} productos en total)")
    return ruta


def procesar_carpeta(carpeta: str = CARPETA_DATOS) -> None:
    xmls = [
        os.path.join(carpeta, nombre)
        for nombre in os.listdir(carpeta)
        if os.path.splitext(nombre)[1].lower() == ".xml"
    ]
    if not xmls:
        print(f"No hay XML en {carpeta}")
        return

    dfs = []
    for ruta_xml in xmls:
        try:
            dfs.append(procesar_xml(ruta_xml, carpeta))
        except Exception as error:
            print(f"[{os.path.basename(ruta_xml)}] ERROR: {error}")

    if dfs:
        df_nuevo = pd.concat(dfs, ignore_index=True)
        guardar_en_base_acumulada(df_nuevo, carpeta)


__all__ = [
    "COLUMNAS_ASPEL",
    "CARPETA_DATOS",
    "NOMBRE_ARCHIVO_RESUMEN",
    "guardar_en_base_acumulada",
    "procesar_carpeta",
    "procesar_xml",
]

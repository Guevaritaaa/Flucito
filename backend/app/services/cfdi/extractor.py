"""Extracción segura de CFDI 4.0 y generación de reportes Excel.

Este módulo no conoce FastAPI. Recibe rutas locales temporales y una ruta
de salida; una ruta HTTP o un agente puede reutilizarlo sin duplicar lógica.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable, Optional
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


NS = {
    "cfdi": "http://www.sat.gob.mx/cfd/4",
    "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital",
}

TIPOS_COMPROBANTE = {
    "I": "Ingreso",
    "E": "Egreso (Nota de Crédito)",
    "T": "Traslado",
    "P": "Complemento de Pago",
    "N": "Nómina",
}

IMPUESTOS = {"001": "ISR", "002": "IVA", "003": "IEPS"}
MAX_XML_BYTES = 20 * 1024 * 1024
CUANTIZADOR = Decimal("0.01")

COLUMNAS_FACTURAS = [
    "UUID", "Archivo_XML", "Version", "Tipo_Codigo", "Tipo", "Serie", "Folio",
    "Fecha_Emision", "Fecha_Timbrado", "RFC_Emisor", "Nombre_Emisor",
    "RegimenFiscal_Emisor", "RFC_Receptor", "Nombre_Receptor",
    "RegimenFiscal_Receptor", "CP_Receptor", "UsoCFDI", "SubTotal", "Descuento",
    "IVA_Trasladado", "Impuestos_Trasladados", "ISR_Retenido", "IVA_Retenido",
    "Impuestos_Retenidos", "Total", "Moneda", "TipoCambio", "Exportacion",
    "MetodoPago", "FormaPago", "CondicionesDePago", "LugarExpedicion",
    "TipoRelacion", "UUID_Relacionados",
]

COLUMNAS_CONCEPTOS = [
    "UUID_Factura", "No", "NoIdentificacion", "ClaveProdServ", "Descripcion",
    "ClaveUnidad", "Unidad", "Cantidad", "ValorUnitario", "Importe", "Descuento",
    "ObjetoImp", "Total_Trasladados", "Total_Retenidos", "Total",
]

COLUMNAS_IMPUESTOS = [
    "UUID_Factura", "No_Partida", "Tipo", "Impuesto_Codigo", "Impuesto",
    "TipoFactor", "TasaOCuota", "Base", "Importe",
]

COLUMNAS_ERRORES = ["Archivo_XML", "Motivo"]
COLUMNAS_MONEDA = {
    "SubTotal", "Descuento", "IVA_Trasladado", "Impuestos_Trasladados",
    "ISR_Retenido", "IVA_Retenido", "Impuestos_Retenidos", "Total",
    "ValorUnitario", "Importe", "Base", "Total_Trasladados", "Total_Retenidos",
}


@dataclass(frozen=True)
class ResultadoExtraccion:
    """Resumen para responder desde API sin leer de nuevo archivo Excel."""

    archivo_excel: Path
    comprobantes: int
    conceptos: int
    impuestos: int
    errores: int


def _texto(elemento: Optional[ET.Element], atributo: str, default: str = "") -> str:
    return elemento.get(atributo, default) if elemento is not None else default


def _decimal(valor: object, default: Decimal = Decimal("0.00")) -> Decimal:
    try:
        if valor is None or str(valor).strip() == "":
            return default
        resultado = Decimal(str(valor))
        return resultado if resultado.is_finite() else default
    except (InvalidOperation, TypeError, ValueError):
        return default


def _importe(valor: object) -> Decimal:
    return _decimal(valor).quantize(CUANTIZADOR, rounding=ROUND_HALF_UP)


def _cargar_xml_seguro(path: Path) -> ET.Element:
    """Bloquea DTD/entidades y XMLs demasiado grandes antes de parsear."""
    if path.stat().st_size > MAX_XML_BYTES:
        raise ValueError(f"XML supera límite de {MAX_XML_BYTES // (1024 * 1024)} MB")

    contenido = path.read_bytes()
    contenido_mayusculas = contenido.upper()
    if b"<!DOCTYPE" in contenido_mayusculas or b"<!ENTITY" in contenido_mayusculas:
        raise ValueError("XML contiene DTD o entidades no permitidas")
    return ET.fromstring(contenido)


def _impuestos_de(elemento: ET.Element, uuid: str, partida: int | str = "") -> list[dict]:
    filas: list[dict] = []
    for tipo, contenedor, etiqueta in (
        ("Traslado", "Traslados", "Traslado"),
        ("Retención", "Retenciones", "Retencion"),
    ):
        ruta = f"./cfdi:Impuestos/cfdi:{contenedor}/cfdi:{etiqueta}"
        for impuesto in elemento.findall(ruta, NS):
            codigo = _texto(impuesto, "Impuesto")
            filas.append({
                "UUID_Factura": uuid,
                "No_Partida": partida,
                "Tipo": tipo,
                "Impuesto_Codigo": codigo,
                "Impuesto": IMPUESTOS.get(codigo, codigo),
                "TipoFactor": _texto(impuesto, "TipoFactor"),
                "TasaOCuota": _texto(impuesto, "TasaOCuota"),
                "Base": _importe(_texto(impuesto, "Base")),
                "Importe": _importe(_texto(impuesto, "Importe")),
            })
    return filas


def _suma_impuestos(filas: list[dict], tipo: str, codigo: Optional[str] = None) -> Decimal:
    return sum(
        (
            _decimal(fila["Importe"])
            for fila in filas
            if fila["Tipo"] == tipo and (codigo is None or fila["Impuesto_Codigo"] == codigo)
        ),
        Decimal("0.00"),
    )


def procesar_cfdi(path: Path) -> dict:
    """Extrae un CFDI 4.0. Lanza ValueError si archivo no se puede usar."""
    try:
        root = _cargar_xml_seguro(path)
    except (ET.ParseError, OSError, ValueError) as error:
        raise ValueError(str(error)) from error

    if root.tag != f"{{{NS['cfdi']}}}Comprobante" or _texto(root, "Version") != "4.0":
        raise ValueError("No es un comprobante CFDI 4.0")

    timbre = root.find(".//tfd:TimbreFiscalDigital", NS)
    emisor = root.find("cfdi:Emisor", NS)
    receptor = root.find("cfdi:Receptor", NS)
    uuid = _texto(timbre, "UUID", "SIN_UUID")
    impuestos_factura = _impuestos_de(root, uuid)

    relacionados = root.findall("./cfdi:CfdiRelacionados", NS)
    tipos_relacion = [_texto(nodo, "TipoRelacion") for nodo in relacionados if _texto(nodo, "TipoRelacion")]
    uuids_relacionados = [
        _texto(relacion, "UUID")
        for nodo in relacionados
        for relacion in nodo.findall("./cfdi:CfdiRelacionado", NS)
        if _texto(relacion, "UUID")
    ]

    factura = {
        "UUID": uuid,
        "Archivo_XML": path.name,
        "Version": _texto(root, "Version"),
        "Tipo_Codigo": _texto(root, "TipoDeComprobante"),
        "Tipo": TIPOS_COMPROBANTE.get(_texto(root, "TipoDeComprobante").upper(), "Desconocido"),
        "Serie": _texto(root, "Serie"),
        "Folio": _texto(root, "Folio"),
        "Fecha_Emision": _texto(root, "Fecha"),
        "Fecha_Timbrado": _texto(timbre, "FechaTimbrado"),
        "RFC_Emisor": _texto(emisor, "Rfc"),
        "Nombre_Emisor": _texto(emisor, "Nombre"),
        "RegimenFiscal_Emisor": _texto(emisor, "RegimenFiscal"),
        "RFC_Receptor": _texto(receptor, "Rfc"),
        "Nombre_Receptor": _texto(receptor, "Nombre"),
        "RegimenFiscal_Receptor": _texto(receptor, "RegimenFiscalReceptor"),
        "CP_Receptor": _texto(receptor, "DomicilioFiscalReceptor"),
        "UsoCFDI": _texto(receptor, "UsoCFDI"),
        "SubTotal": _importe(_texto(root, "SubTotal")),
        "Descuento": _importe(_texto(root, "Descuento")),
        "IVA_Trasladado": _importe(_suma_impuestos(impuestos_factura, "Traslado", "002")),
        "Impuestos_Trasladados": _importe(_suma_impuestos(impuestos_factura, "Traslado")),
        "ISR_Retenido": _importe(_suma_impuestos(impuestos_factura, "Retención", "001")),
        "IVA_Retenido": _importe(_suma_impuestos(impuestos_factura, "Retención", "002")),
        "Impuestos_Retenidos": _importe(_suma_impuestos(impuestos_factura, "Retención")),
        "Total": _importe(_texto(root, "Total")),
        "Moneda": _texto(root, "Moneda", "MXN"),
        "TipoCambio": _texto(root, "TipoCambio"),
        "Exportacion": _texto(root, "Exportacion"),
        "MetodoPago": _texto(root, "MetodoPago"),
        "FormaPago": _texto(root, "FormaPago"),
        "CondicionesDePago": _texto(root, "CondicionesDePago"),
        "LugarExpedicion": _texto(root, "LugarExpedicion"),
        "TipoRelacion": ", ".join(tipos_relacion),
        "UUID_Relacionados": ", ".join(uuids_relacionados),
    }

    conceptos: list[dict] = []
    impuestos_conceptos: list[dict] = []
    for numero, concepto in enumerate(root.findall("./cfdi:Conceptos/cfdi:Concepto", NS), start=1):
        impuestos = _impuestos_de(concepto, uuid, numero)
        impuestos_conceptos.extend(impuestos)
        importe = _decimal(_texto(concepto, "Importe"))
        descuento = _decimal(_texto(concepto, "Descuento"))
        traslados = _suma_impuestos(impuestos, "Traslado")
        retenciones = _suma_impuestos(impuestos, "Retención")
        conceptos.append({
            "UUID_Factura": uuid,
            "No": numero,
            "NoIdentificacion": _texto(concepto, "NoIdentificacion"),
            "ClaveProdServ": _texto(concepto, "ClaveProdServ"),
            "Descripcion": _texto(concepto, "Descripcion"),
            "ClaveUnidad": _texto(concepto, "ClaveUnidad"),
            "Unidad": _texto(concepto, "Unidad"),
            "Cantidad": _decimal(_texto(concepto, "Cantidad")),
            "ValorUnitario": _importe(_texto(concepto, "ValorUnitario")),
            "Importe": _importe(importe),
            "Descuento": _importe(descuento),
            "ObjetoImp": _texto(concepto, "ObjetoImp"),
            "Total_Trasladados": _importe(traslados),
            "Total_Retenidos": _importe(retenciones),
            "Total": _importe(importe - descuento + traslados - retenciones),
        })

    return {"factura": factura, "conceptos": conceptos, "impuestos": impuestos_conceptos}


def _crear_hoja(libro: Workbook, nombre: str, columnas: list[str], filas: list[dict]) -> None:
    hoja = libro.create_sheet(nombre)
    hoja.append(columnas)
    for fila in filas:
        hoja.append([
            float(valor) if isinstance(valor, Decimal) else valor
            for valor in (fila.get(columna, "") for columna in columnas)
        ])

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions
    for columna, celda in enumerate(hoja[1], start=1):
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F4E78")
        encabezado = str(celda.value)
        ancho = max(12, min(45, len(encabezado) + 2))
        for fila in range(2, min(hoja.max_row, 200) + 1):
            valor = hoja.cell(fila, columna).value
            if valor is not None:
                ancho = max(ancho, min(45, len(str(valor)) + 2))
        if encabezado in COLUMNAS_MONEDA:
            for fila in range(2, hoja.max_row + 1):
                hoja.cell(fila, columna).number_format = "#,##0.00"
        hoja.column_dimensions[get_column_letter(columna)].width = ancho


def generar_excel_desde_xmls(rutas_xml: Iterable[Path], ruta_salida: Path) -> ResultadoExtraccion:
    """Procesa XMLs y crea Excel con Facturas, Conceptos, Impuestos y Errores."""
    facturas: list[dict] = []
    conceptos: list[dict] = []
    impuestos: list[dict] = []
    errores: list[dict] = []
    uuids_vistos: set[str] = set()

    for ruta_xml in rutas_xml:
        try:
            datos = procesar_cfdi(ruta_xml)
            uuid = datos["factura"]["UUID"]
            if uuid != "SIN_UUID" and uuid in uuids_vistos:
                raise ValueError(f"UUID duplicado: {uuid}")
            if uuid != "SIN_UUID":
                uuids_vistos.add(uuid)
            facturas.append(datos["factura"])
            conceptos.extend(datos["conceptos"])
            impuestos.extend(datos["impuestos"])
        except ValueError as error:
            errores.append({"Archivo_XML": ruta_xml.name, "Motivo": str(error)})

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    libro = Workbook()
    libro.remove(libro.active)
    _crear_hoja(libro, "Facturas", COLUMNAS_FACTURAS, facturas)
    _crear_hoja(libro, "Conceptos", COLUMNAS_CONCEPTOS, conceptos)
    _crear_hoja(libro, "Impuestos", COLUMNAS_IMPUESTOS, impuestos)
    _crear_hoja(libro, "Errores", COLUMNAS_ERRORES, errores)
    libro.save(ruta_salida)

    return ResultadoExtraccion(
        archivo_excel=ruta_salida,
        comprobantes=len(facturas),
        conceptos=len(conceptos),
        impuestos=len(impuestos),
        errores=len(errores),
    )

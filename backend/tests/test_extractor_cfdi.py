from pathlib import Path
from decimal import Decimal

from openpyxl import Workbook
from openpyxl import load_workbook

from app.services.cfdi.extractor import _crear_hoja, generar_excel_desde_xmls


BASE_DIR = Path(__file__).parent
XML_VALIDO = BASE_DIR / "fixtures" / "cfdi_valido.xml"


def test_generar_excel_desde_cfdi_valido(tmp_path: Path) -> None:
    salida = tmp_path / "facturas.xlsx"

    resultado = generar_excel_desde_xmls([XML_VALIDO], salida)

    assert salida.is_file()
    assert resultado.comprobantes == 1
    assert resultado.conceptos == 1
    assert resultado.impuestos == 2
    assert resultado.errores == 0

    libro = load_workbook(salida, data_only=True)
    try:
        assert libro.sheetnames == ["Facturas", "Conceptos", "Impuestos", "Errores"]

        factura = libro["Facturas"]
        encabezados = {celda.value: indice for indice, celda in enumerate(factura[1], start=1)}
        assert factura.cell(2, encabezados["UUID"]).value == "123e4567-e89b-12d3-a456-426614174000"
        assert factura.cell(2, encabezados["IVA_Trasladado"]).value == 16
        assert factura.cell(2, encabezados["ISR_Retenido"]).value == 8
        assert factura.cell(2, encabezados["Total"]).value == 108
    finally:
        libro.close()


def test_formato_monetario_cubre_mas_de_200_filas() -> None:
    libro = Workbook()
    libro.remove(libro.active)

    _crear_hoja(
        libro,
        "Conceptos",
        ["Importe"],
        [{"Importe": Decimal("1.00")} for _ in range(201)],
    )

    assert libro["Conceptos"]["A202"].number_format == "#,##0.00"
